import openpyxl

class ExcelMethods_class:

    def __init__(self, file):
        # File ek hi baar khulegi — constructor me
        self.file = file
        self.workbook = openpyxl.load_workbook(file)

    def getRowCount(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet.max_row

    def read_data(self, sheetName, rownum, columnno):
        sheet = self.workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def write_data(self, sheetName, rownum, columnno, data):
        sheet = self.workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        self.workbook.save(self.file)
        
